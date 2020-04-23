import json
import jsonpath
import requests
import openpyxl

class Common:

    def __init__(self, FileNamePath, SheetName):
        # Excel code
        global workbook
        global sheet
        workbook = openpyxl.load_workbook(FileNamePath)
        sheet = workbook[SheetName]

    def fetch_row_count(self):
        rows = sheet.max_row
        return rows

    def fetch_column_count(self):
        cols = sheet.max_column
        return cols

    def fetch_key_names(self):
        c = sheet.max_column
        names = []

        for i in range(1,c+1):
            key_name = sheet.cell(row=1,column=i)
            names.insert(i-1, key_name.value)

        return names

    def update_request_with_data(self, rowNumber, jsonRequest, keyList):
        c = sheet.max_column
        for i in range(1, c+1):
            name = sheet.cell(row=rowNumber, column=i)
            jsonRequest[keyList[i-1]] = name.value


        return jsonRequest
