import requests
import openpyxl
import json
import jsonpath

def test_add_multiple_students():
    #API
    api_url = "http://thetestingworldapi.com/api/studentsDetails"
    f = open('C:\\Users\\Harshita Sahay\\PycharmProjects\\RestAPIAutomation\\AddNewStudent.json')

    #Excel code
    workbook = openpyxl.load_workbook('C:\\Users\\Harshita Sahay\\PycharmProjects\\RestAPIAutomation\\TestMultipleData.xlsx')
    sheet = workbook['Sheet1']
    rows = sheet.max_row
    json_request = json.loads(f.read())

    for i in range(2, rows+1):
        cell_first_name = sheet.cell(row = i, column=1)
        cell_middle_name = sheet.cell(row=i, column=2)
        cell_last_name = sheet.cell(row=i, column=3)
        cell_dob = sheet.cell(row=i, column=4)

        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_middle_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_dob.value

        response = requests.post(api_url, json_request)

        print(response.text)
        print(response.status_code)
        assert response.status_code == 201
